#coding:utf8

"""
Author : zhangwenqi
Description : 读取excel内容
CreateTime : 2019-08-05
"""
import openpyxl

class ObtainEexcel():
    def __init__(self,filename,sheetname):
        self.filename = filename
        self.handxl = openpyxl.load_workbook(filename)
        self.sheet = self.handxl.get_sheet_by_name(sheetname)
        self.row_no = self.sheet.max_row
        self.column_no = self.sheet.max_column

    """获取所有的sheet名称"""
    def obtain_sheets_name(self):
        return self.handxl.get_sheet_names()

    """获取表格内容"""
    def obtain_table_value(self,row,column):
        return self.sheet.cell(row=row,column=column).value

    """写入表格内容"""
    def set_table_value(self, row, column,value):
        self.sheet.cell(row=row, column=column).value = value

    """获取所有行数据"""
    def obtain_table_rows_value(self):
        list_rows_value = []
        for i in self.sheet.iter_rows():
            list_rows_value.append([j.value for j in i])
        return list_rows_value

    """获取所有列数据"""
    def obtain_table_columns_value(self):
        list_columns_value = []
        for i in self.sheet.iter_cols():
            list_columns_value.append([j.value for j in i])
        return list_columns_value

    """通过表名标题去找对应数据"""
    def obtain_position_by_table(self,table_name):
        tables = self.sheet.iter_rows().__next__()
        return [i.value for i in tables].index(table_name)+1

    def obtain_values_by_position(self,table_name,target_value):
        position = self.obtain_position_by_table(table_name)
        list_values = self.obtain_table_rows_value()
        result_values = []
        for i in list_values:
            if i[position-1] == target_value:
                result_values.append(i)
        return result_values

    def table_values_sort(self,table_name,values):
        position = self.obtain_position_by_table(table_name)
        return sorted(values,key=lambda x:x[position-1])


    """保存excel"""
    def save(self):
        self.handxl.save(self.filename)
        self.handxl.close()