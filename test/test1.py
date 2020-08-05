import openpyxl

def func(s):
    for i in s:
        print("%s:%s" % (i,s[i]))


if __name__ == "__main__":
    s = {"sign":"44549876042bc0cce5f57fd1432e6b92","company_id":"8968336369328084993","out_order_id":"21666097","callback_status":"10001","order_id":"611986578947567258","order_source":"1","new_order":"{\"order_id\":\"14110766\"}","msg":"","test_function":"OrderStatusCallbackTest_testAppRestartOrder#"}

    func(s)
    # file = r"/Users/didi/Desktop/111.xlsx"
    # wb = openpyxl.load_workbook(file)
    # wb.create_sheet("3")
    # d = {}
    # st = wb["1"]
    # n = st.max_row
    # for i in range(2,n+1):
    #     v = (lambda x: x.strip() if x else x)(st["A%s" % i].value)
    #     if v:
    #         if v not in d:
    #             d[v] = {"j1":st["B%s" % i].value,"b":"1","j2":"","j3":""}
    #         else:
    #             if d[v]["j1"] != st["B%s" % i].value:
    #                 d[v]["j3"] = st["B%s" % i].value
    #                 d[v]["b"] = "2"
    # st = wb["2"]
    # n = st.max_row
    # for i in range(2,n+1):
    #     v = (lambda x: x.strip() if x else x)(st["A%s" % i].value)
    #     if v:
    #         if v not in d:
    #             d[v] = {"j1":st["B%s" % i].value,"b":"1","j2":"","j3":""}
    #         else:
    #             if d[v]["j1"] != st["B%s" % i].value:
    #                 d[v]["j2"] = st["B%s" % i].value
    #                 d[v]["b"] = "2"
    # st = wb["3"]
    # n = 2
    # for i in d:
    #     st["A%s" % n] = i
    #     st["B%s" % n] = d[i]["j1"]
    #     st["C%s" % n] = d[i]["j2"]
    #     st["D%s" % n] = d[i]["j3"]
    #     st["E%s" % n] = d[i]["b"]
    #     n += 1
    # wb.save(file)